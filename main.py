import os
import sys
import pandas as pd
import random
import util
import ast
from pydantic import BaseModel, create_model, Field
from eval import get_metric_sums, get_metric_dist 
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

#new
import shutil
from openpyxl import load_workbook

from datetime import datetime

# Togglable Options
IS_DEBUG = False # If True, additional debug statements will be printed.
IS_SKIP_DATA_PROCESS = False # If True, skip 'create_args_sheet' - Use if the program crashes after the data has been processed
TOTAL_SESSIONS = 4 # The highest numbered session in the data
IS_ANALYZE_ALL_SESSIONS = False # If True, all sessions will be analyzed. If False, only 'session_num' will be analyzed.
session_num = 'test3' # The single session to analyze if 'IS_ANALYZE_ALL_SESSIONS' is False
# IMPORTANT: One, and only one, of the three booleans below should be True at all times.
IS_DOWNLOADED_CSV = False # Set to True if the data being used is from the "Step 1: Downloaded CSVs" Google Drive
IS_READY_FOR_FILEREAD = True # Set to True if the data being used is from the "Step 2: Ready for Fileread Download" Google Drive
IS_FILEREAD_RESULTS = False # Set to True if the data being used is from the "Step 3: Fileread Results" Google Drive

DATA_DIR = 'data'
PROCESSING_DIR = 'processing'
RESULTS_DIR = 'results'
sys.path.append(os.getcwd())

# collecting arguments in a deliberation
def extract_args(path, deliberation):
    df = pd.DataFrame(pd.read_excel(path)) 
    start_col = df.columns.get_loc("All Arguments Summarized")
    cols = df.columns[start_col:]

    # filtering out debugging columns
    filter_cols = [col for col in cols if col.find("debugging") == -1]
    relevant_args_df = df.dropna(subset=filter_cols, how='all')

    # list of Fileread's summarized arguments and their indexed order
    arguments = list((relevant_args_df["All Arguments Summarized"], relevant_args_df["Order"]))
    arguments = [(x, y) for x, y in zip(arguments[0], arguments[1]) if str(x) != 'nan']
    return concat_args(arguments, deliberation)

# cleaning arguments
def concat_args(arguments, deliberation):
  cleaned_args = {deliberation: []}
  for arg in arguments:
    args = arg[0].split("\n")
    joined = " ".join([arg[3:] for arg in args])
    cleaned_args[deliberation].append((joined, arg[1]))
  return cleaned_args

# Given a subset of arguments,
# return a list of topics that most arguments deliberate on.
def extract_topics(sampled_args, attempts = 0):
   NUM_TOPICS = '7'
   if attempts == 0: print("\nGenerating", NUM_TOPICS, "topics from Session", session_num + "...")

   prompt = """This is a list of arguments presented in a deliberation. Your job is to identify the single, primary topic deliberated on and write
   """ + NUM_TOPICS + """ distinct policies regarding the primary topic in a Python list of strings, with each string being a policy.
   The first string in the list is the primary topic, and every other string is a policy.
   It should be possible to define every argument initially provided as being an argument "for" or "against" at least one of the policies you generate.
   The exceptions to this are the very small number of arguments that are unrelated to the primary topic and all policies.
   Here is an example of what you might return if the primary topic was 'Ranked choice voting':
   ["Ranked choice voting", "Implement ranked choice voting as an alternative method both to elected officials and representatives at all levels", "Use proportional representatives to elect elected officials"...]
   The policies in the list should be somewhat distinct from each other.
   You should NOT have broad policies, such as the advantages and disadvantages of the primary topic. Rather,
   you should instead find nuanced groups of arguments that may be present on either side of the discussion.
   Every policy should be a feasable, actionable change. Here is an example of what a policy should NOT be:
   "Limit the influence of political parties in elections" This policy is BAD because it is too vague and not actionable;
   this BAD policy could not feasibly be turned into a law.
   The following non-case-sensitive words are BANNED and should NOT be included in your response:
   "and", "but", "or", "by", "Promote", "Encourage", "Ensure", "Explore", "Maintain".
   This program will CRASH if your response fails to conform to these guidelines.
   Other code ran for several hours before you were given this task.
   If this program crashes, we will need to restart which takes several hours.
   Please take your time and ensure ALL of these instructions are followed.
   Do NOT write "Primary topic:" in your response.
   Do NOT number the list.
   Do NOT indent in your response.
   Do NOT include a newline character in your response.
   Do NOT include anything in the list other than the primary topic and the policies themselves.
   Your response is ONLY a single list of the primary topic and the policies.
   Once again, your response is ONLY a single list.
   Your response is ONLY one line.
   VERY IMPORTANT: EVERY argument initially provided to you should be an argument "for" or "against" at least one of the policies you generate.
   Once again, EVERY argument should be relevant to one of the policies you return.
   Please, the Python list you return should contain EXACTLY """ + str(int(NUM_TOPICS)+1) + """ strings.
   Thank you!"""
   response = util.simple_llm_call(prompt, sampled_args)
   if IS_DEBUG: print("\n(DEBUG) Raw response:", response + "\n")
   topic_list = ast.literal_eval(response.strip())
   # If invalid response, try again
   if len(topic_list) != (int(NUM_TOPICS) + 1) or type(topic_list) != list:
     attempts += 1
     print("Failed attempt #" + str(attempts))
     # If failed too many times, give up
     if attempts >= 5:
       error("Failed to generate a primary topic and policies")
     # Else, try again
     print("Retrying...")
     return extract_topics(sampled_args, attempts)
   # Else, success!
   print_topics(topic_list)
   return topic_list

# Given a list of topics,
# print them.
def print_topics(topic_list):
  print("Primary Topic:", topic_list[0])
  for i in range(1, len(topic_list)):
    print("Policy", str(i) + ":", topic_list[i])

# Given a list of policies,
# return a list of variables that can be used as shorthand for each policy.
def generate_policy_variables(topics, attempts = 0):
   NUM_VARIABLES = str(len(topics) - 1)
   PRIMARY_TOPIC = topics[0]

   policy_list = []
   for i in range(int(NUM_VARIABLES)):
      policy_list.append(topics[i + 1])

   if attempts == 0: print("Generating variables for the", NUM_VARIABLES, "policies regarding", PRIMARY_TOPIC + "...")
   
   prompt = """This is a list of policies representing arguments made in a deliberation about """ + PRIMARY_TOPIC + """.
   Your job is to create variable names for each policy in a Python list of strings, with each string being a variable name.
   Your response will be """ + NUM_VARIABLES + """ strings.
   The variable names should be short.
   The variable names should use camel case style.
   The variable names should encapsulate the core of the policy.
   Here is an example of what you might return:
   ["VariableOne", "VariableTwo",...]
   This program will CRASH if your response fails to conform to these guidelines.
   Other code ran for several hours before you were given this task.
   If this program crashes, we will need to restart which takes several hours.
   Please take your time and ensure ALL of these instructions are followed.
   Do NOT number the list.
   Do NOT indent in your response.
   Do NOT include a newline character in your response.
   Do NOT include anything in the list other than the variable names.
   Your response is ONLY a single list of variable names.
   Once again, your response is ONLY a single list.
   Your response is ONLY one line.
   One again, the variable names must be SHORT
   The Python list you return should contain EXACTLY """ + NUM_VARIABLES + """ strings.
   Thank you!"""
   response = util.simple_llm_call(prompt, policy_list)
   if IS_DEBUG: print("\n(DEBUG) Raw response:", response + "\n")
   response_list = ast.literal_eval(response.strip())
   # If invalid response, try again
   if len(response_list) != int(NUM_VARIABLES) or type(response_list) != list:
     attempts += 1
     print("Failed attempt #" + str(attempts))
     # If failed too many times, give up
     if attempts >= 5:
       error("Failed to generate variables for each policy")
     # Else, try again
     print("Retrying...")
     return generate_policy_variables(topics, attempts)
   # Else, success!
   print_list(response_list, "Variable")

   # Generate results, metrics, and key
   results_path = os.path.join(RESULTS_DIR, session_num)
   create_results_path(results_path)
   generate_key(topics, response_list, results_path)
   
   variable_list = []
   for i in range(len(response_list)):
      variable_list.append("FOR: " + response_list[i])
      variable_list.append("AGAINST: " + response_list[i])
   variable_list.append("other")
   variable_list.append("notRelevant")
   if IS_DEBUG:
      print("\n(DEBUG) variable_list:")
      print_list(variable_list)
   return variable_list

# Given a list and (optionally) a header,
# print it
def print_list(list, header = ""):
   for i in range(len(list)):
      print(header, str(i) + ":", list[i])

# Given a list of categories,
# construct a JSON class for the model to use.
def build_JSON_class(category_variables):
  print("\nBuilding JSON class for the model...")
  fields = {var: (bool, Field(default=False)) for var in category_variables}
  Categories = create_model("Categories", **fields)
  if IS_DEBUG:
    categories_instance = Categories()
    print("\n(DEBUG) JSON Categories:\n" + categories_instance.model_dump_json(indent=2) + "\n")
  return Categories()
  
# Given a primary topic and specific policies,
# return a prompt that the model will use to judge arguments.
def build_argument_analysis_prompt(topics, policy_variables):
  NUM_POLICIES = len(topics) - 1
  PRIMARY_TOPIC = topics[0]

  policies = []
  for_policies = []
  against_policies = []
  for i in range(NUM_POLICIES):
     policies.append(topics[i + 1])
     for_policies.append(policy_variables[i * 2])
     against_policies.append(policy_variables[i * 2 + 1])

  policy_instructions = """"""
  for i in range(NUM_POLICIES):
     policy_instructions += """\nReturn true for """ + for_policies[i] + """ if the argument is IN SUPPORT OF """ + policies[i] + """.\nReturn false for """ + for_policies[i] + """ otherwise.
     Return true for """ + against_policies[i] + """ if the argument is IN OPPOSITION TO """ + policies[i] + """\nReturn false for """ + against_policies[i] + """ otherwise.\n"""
  
  prompt = """You are a skilled annotator tasked with identifying the type of arguments made in a deliberation about """ + PRIMARY_TOPIC + """.
          Read through the given arguments presented.
          Your job is to extract the type of arguments made and load them into the JSON object containing true/false parameters.
          """ + policy_instructions + """
          Return true for other if the argument is relevant to """ + PRIMARY_TOPIC + """
          but not covered by the other categories.

          Return true for notRelevant if the argument is not relevant to the discussion of """ + PRIMARY_TOPIC + """.

          It is possible for a single argument to have multiple true parameters, except for notRelevant.
          If an argument is not relevant, all other parameters should be false.
          You should return true for at least one parameter."""
  
  if IS_DEBUG: print("\n(DEBUG) build_argument_analysis_prompt Prompt:\n" + prompt + "\n")
  
  return prompt

# adds an LLM's topic classifications for an argument to the deliberation df
def add_results(response, df, line):
   for key in response.keys():
      df.loc[df['Order'] == line, key] = response[key]

# classifies all arguments in all deliberations based on the extracted topics
# note: most time-expensive function to call / may need to increase token size
def arg_inference(all_args_indexed, results_path, argument_analysis_prompt, json):
  print("\nAnalyzing Session", session_num, "deliberations...")
  # looping over all deliberations
  for deliberation in all_args_indexed.keys():
    args = all_args_indexed[deliberation]
    path = os.path.join(PROCESSING_DIR, session_num, deliberation)

    # initializing df and fields
    df = pd.DataFrame(pd.read_excel(path)) 
    for key in json.model_fields.keys():
      df[key] = False

    # loop over a deliberation's arguments
    for arg in args:
      prompt = argument_analysis_prompt
      response = util.json_llm_call(prompt, arg[0], json)
      line = arg[1]
      add_results(response, df, line)
    new_filename = "EVALUATED" + deliberation.replace("xlsx", "csv")
    df.to_csv(os.path.join(results_path, new_filename), index=False)
    print("Created", new_filename)
  print("Finished analyzing deliberations in Session", session_num)

# Given a path for a 'results' folder,
# create that folder and a 'metrics' subfolder if needed.
def create_results_path(results_path):
  print("\nCreating Session", session_num, "results and metrics folders...")
  # Creates the required results folder if it does not already exist
  os.makedirs(results_path, exist_ok=True)

  # Creates the required metrics folder if it does not already exist
  metrics_path = os.path.join(results_path, "metrics")
  os.makedirs(metrics_path, exist_ok=True)

# Given a path for a 'processing' folder,
# create that folder if needed.
def create_processing_path(processing_path):
   print("\nCreating Session", session_num, "processing folder...")
   # Creates the required results folder if it does not already exist
   os.makedirs(processing_path, exist_ok=True)

# this function takes in text from a deliberation and determines if it contains arguments
def check_arguments(text):
    system_prompt = "Determine if the following text contains arguments. Respond ONLY with 'yes' or 'no'. Do not provide an explanation, do not capitalize Yes or No, do not put any punctuation, only respond 'yes' or 'no'."
    response = util.simple_llm_call(system_prompt, text)
    # if invalid response, try again
    if response not in ['yes', 'no']:
        # print("bad output")
        response = util.simple_llm_call(system_prompt, text)
    # print(text + "----->" + response)
    return response.strip().lower() == 'yes'

# Checks Togglable Options,
# throws an error if they are invalid.
def valid_options_check():
  error_message = "The Togglable Options have been incorrectly set up. Please refer to 'Togglable Options' at the top of main.py and ensure everything is set up correctly. Currently, the data is marked as coming from multiple sources which is not allowed. Please review 'IS_DOWNLOADED_CSV', 'IS_READY_FOR_FILEREAD', and 'IS_FILEREAD_RESULTS' for more information."
  if IS_DOWNLOADED_CSV and IS_READY_FOR_FILEREAD:
    error(error_message)
  elif IS_DOWNLOADED_CSV and IS_FILEREAD_RESULTS:
    error(error_message)
  elif IS_READY_FOR_FILEREAD and IS_FILEREAD_RESULTS:
    error(error_message)

# Given a sheet,
# format it appropriately.
def format_sheet(sheet_path):
  valid_options_check() # Ensures all Togglable Options are valid

  wb = load_workbook(sheet_path)
  ws = wb.worksheets[0]
  
  if IS_DOWNLOADED_CSV:
     ws.delete_rows(1)
     ws.delete_cols(1)
     ws.delete_cols(2, 2)
     ws.cell(row=1, column=1).value = "speaker"
     ws.cell(row=1, column=3).value = "text"

  wb.save(sheet_path)

# Given a data folder,
# clean the data in it.
def clean_data(data_path):
   for deliberation in os.listdir(data_path):
      csv_to_xlsx(data_path, deliberation)
   for deliberation in os.listdir(data_path):
      path = os.path.join(data_path, deliberation)
      if path.endswith('xlsx'):
        wb = load_workbook(path)
        ws = wb.worksheets[0]
        ws.title = "in"
        for sheet_name in wb.sheetnames:
            if sheet_name != "in":
              wb.remove(wb[sheet_name])
    
        wb.save(path)

# this function duplicates the input spreadsheet in destination folder, and creates and populates contians args column
def create_args_sheet(file_path, destination_folder):
    df = pd.read_excel(file_path)
    file_name = os.path.basename(file_path)
    duplicated_file_path = os.path.join(destination_folder, file_name)

    # Duplicate the file to the destination folder with the new name
    shutil.copy(file_path, duplicated_file_path)

    format_sheet(duplicated_file_path) # Format the file

    # Load the duplicated file with openpyxl to add columns
    wb = load_workbook(duplicated_file_path)
    for sheet in wb.sheetnames:
       ws = wb[sheet]
    
    # Define columns
    ORDER_COL = 1 # Order
    TEXT_COL = 4 # text
    HAS_ARG_COL = 5 # Has Arguments
    ARG_SUM_COL = 6 # All Arguments Summarized

    # Add the new columns
    ws.insert_cols(1) # Insert "Order" column
    ws.cell(row=1, column=ORDER_COL, value="Order")
    ws.cell(row=1, column=HAS_ARG_COL, value="Has Arguments")
    ws.cell(row=1, column=ARG_SUM_COL, value="All Arguments Summarized")

    # Iterate over the rows and populate the "has arguments" column
    for row in range(2, ws.max_row + 1):
        text = ws.cell(row=row, column=TEXT_COL).value
        if text is not None:  # Check if the cell is not empty
            has_arguments = check_arguments(text)
            ws.cell(row=row, column=HAS_ARG_COL, value="yes" if has_arguments else "no")
        else:
            ws.cell(row=row, column=HAS_ARG_COL, value="no")
        # Construct the "Order" column
        ws.cell(row=row, column=ORDER_COL, value=row-1)

    # Iterate over the rows again to populate the "all arguments summarized" column
    for row in range(2, ws.max_row + 1):
        has_arguments = ws.cell(row=row, column=HAS_ARG_COL).value
        if has_arguments == "yes":
            text = ws.cell(row=row, column=TEXT_COL).value
            # summarized_text = util.simple_llm_call("Summarize the following text in a numbered list. For example, the text input ' thank you. so the last thing was that if there's ten candidates and you only pick your top two, then your and they don't get it, your vote doesn't count it all. so the best thing to do is you got to rank all 10, but again, not enough data on the subject to make predictions. and i think mike use' should be summarized as '1. If you only pick your top two candidates out of ten, your vote doesn't count if they don't win. 2. The best approach is to rank all ten candidates. 3. There is not enough data available to make predictions on the subject.' Follow this formatting exactly.", text)
            summarized_text = util.simple_llm_call("Summarize the following text in a numbered list.  Follow the output format '1. argument 1 \n 2. argument 2 \n' etcetera. Every argument MUST be on a different line. There could be one or more arguments.", text)
            ws.cell(row=row, column=ARG_SUM_COL).value = summarized_text
            #print("+++++++++" + text)
            #print("==========" + summarized_text)

    # Save the modified duplicated file
    print("Processed", file_name)
    wb.save(duplicated_file_path)

# Called when an error occurs
def error(reason = "No reason provided"):
   print("\n\nERROR:", reason)
   print("\nProgram Terminated\n")
   sys.exit()

# Given a deliberation,
# convert it from a CSV file into a XLSX file if needed.
def csv_to_xlsx(data_path, deliberation):
  path = os.path.join(data_path, deliberation)
  if path.endswith('numbers'): error("""".numbers" files are not supported. Please manually convert them to ".csv" files using Apple's Numbers Application.\n1. Open the .numbers file in Apple Numbers.\n2. Go to File -> Export To -> CSV.\n3. Save the file as a ".csv".""")
  if path.endswith('csv'):
    df = pd.read_csv(path, header=1)

    # Convert Headers
    headers = pd.DataFrame([df.columns.tolist()], columns=df.columns[:6]) # Get column headers
    df = pd.concat([headers, df]) # Add column headers to the top of the df

    # Convert Room ID
    with open(path, 'r') as f:
      room_ID_list = f.readline().strip().split(',') # Get a list of first row values
    room_ID_row = pd.DataFrame([room_ID_list], columns=df.columns[:2]) # Convert to a df row
    room_ID_row.loc[0, "userId"] = pd.to_numeric(room_ID_row.loc[0, "userId"]) # Convert roomID number from text to a numeric value
    df = pd.concat([room_ID_row, df]) # Add Room ID row to the top of the df

    # Convert from CSV to XLSX
    new_file = deliberation.replace('csv', 'xlsx')
    new_path = os.path.join(data_path, new_file)
    df.to_excel(new_path, index=False, header=False)

    os.remove(path)

# Upades excel file with "for", "against", "other", and "not relevant" columns
def format_excel(path, topics):
   # Read the existing Excel file
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Convert worksheet to DataFrame
    df = pd.DataFrame(ws.values)

    # Set the first row as column headers
    df.columns = df.iloc[0]
    df = df[1:]

    # Start adding new columns from column 7
    start_col = 6

    # Add columns for each policy (up to 7)
    for i, policy in enumerate(topics[1:8], start=1):  # Skip the first item (primary topic)
        df.insert(start_col, f"for: {policy}", "")
        df.insert(start_col + 1, f"against: {policy}", "")
        start_col += 2

    # Add "other" and "not relevant" columns
    df.insert(start_col, "other", "")
    df.insert(start_col + 1, "not relevant", "")

    # Clear the worksheet and write the updated DataFrame
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
  
    # Save the changes
    wb.save(path)
    print(f"Updated {path} with new policy columns")
  
# adds the classified argument to the Excel sheet
def add_arg_result(int_response, line, path, arg):
   # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    # Calculate the actual row number (add 1 because Excel is 1-indexed)
    actual_row = line + 1

    # Get the current cell value
    current_value = sheet.cell(row=actual_row, column=int_response).value

    # If there's already content, append the new arg. Otherwise, use the new arg.
    if current_value:
        new_value = f"{current_value} {arg}"
    else:
        new_value = arg

    # Update the cell with the new value
    sheet.cell(row=actual_row, column=int_response, value=new_value)

    # Save the workbook
    wb.save(path)

    print(f"Added '{arg}' to row {actual_row}, column {int_response}")

#cleans up output repsponse by stopping before first newline character
def response_clean(response):
  newline_index = response.find('\n')
  if newline_index != -1:
    response = response[:newline_index]
  backslash_index = response.find('\\')
  # If a backslash is found, return the substring up to the backslash
  if backslash_index != -1:
    response = response[:backslash_index]
  return response.replace(" ", "")

# classifies all arguments in all deliberations in session_num based on the generated topics
# note: time-expensive
def arg_sort(all_args_indexed, topics):
  print("\nAnalyzing Session", session_num, "deliberations...")

  prompt = """You are a skilled annotator tasked with catagorizing the type of arguments made in a deliberation about """ + topics[0] + """.
            Read through the given argument presented.
            If the argument is in favor of """ + topics[1] + """, return "7".
            If the argument is against """ + topics[1] + """, return "8".
            If the argument is in favor of """ + topics[2] + """, return "9".
            If the argument is against """ + topics[2] + """, return "10".
            If the argument is in favor of """ + topics[3] + """, return "11".
            If the argument is against """ + topics[3] + """, return "12".
            If the argument is in favor of """ + topics[4] + """, return "13".
            If the argument is against """ + topics[4] + """, return "14".
            If the argument is in favor of """ + topics[5] + """, return "15".
            If the argument is against """ + topics[5] + """, return "16".
            If the argument is in favor of """ + topics[6] + """, return "17".
            If the argument is against """ + topics[6] + """, return "18".
            If the argument is in favor of """ + topics[7] + """, return "19".
            If the argument is against """ + topics[7] + """, return "20".
            If the argument discusses a category that is relevant to """ + topics[0] + """ but not covered by the other categories, return '21'.
            If the argument is not relevant to the discussion of """ + topics[0] + """, return "22".
            Only return one of these number options.  Do not include punctuation or any words except for the number.  Do not add extra text after the answer.  Your response should only be one or two characters depening if the answer is a one or two digit number.  Do not put a space before the answer. Create the shortest possible response.
            """
  print(prompt)

  # looping over all deliberations
  for deliberation in all_args_indexed.keys():
    args = all_args_indexed[deliberation]
    path = os.path.join(PROCESSING_DIR, session_num, deliberation)

    # add columns to Excel sheets to prepare for classification
    format_excel(path, topics)

    # loop over a deliberation's arguments
    for arg_group in args:
      line = arg_group[1] # zero based indexed (eg 9 means line 10 of the Excel sheet)
      arg_group_parsed = arg_group[0].split(".")
      for arg in arg_group_parsed:
        if arg == "": # skip empty strings
           continue
        response = util.simple_llm_call(prompt, arg)
        response = response_clean(response)
        counter = 0
        print("RESPONSE:" + response)
        while response not in [str(i) for i in range(7, 23)]:
           response = util.simple_llm_call(prompt, arg)
           response = response_clean(response)
           counter += 1
           if counter >= 5: # couter to prevent infinite loops
              response = "22"
              break
        int_response = int(response)
        add_arg_result(int_response, line, path, arg)

def generate_key(topics, policy_variables, results_path):
   KEY_NAME = "KEY_Session_" + session_num + ".txt"
   key_path = os.path.join(results_path, "metrics", KEY_NAME)

   key_text = "Session " + session_num + " Key"
   time = datetime.now().strftime("%Y-%m-%d at %H:%M:%S")
   key_text += "\nGenerated on: " + time
   key_text += "\n\nPrimary Topic: " + topics[0]
   key_text += "\n\nPolicy Key:"
   for i in range(len(policy_variables)):
      key_text += "\n* " + policy_variables[i] + " = " + topics[i+1]
   
   if IS_DEBUG: print("\n(DEBUG) key_text:\n" + key_text)
   
   with open(key_path, 'w') as key:
      key.write(key_text)

def main():
    print("Entering main() of Session", session_num)

    # keys = deliberation ids, values = (argument, index in deliberation)
    all_args_indexed = {}
    all_args = []

    # looping over all deliberations and collecting 1) the arguments presented and 2) the index of each argument in that deliberation
    data_path = os.path.join(DATA_DIR, session_num)
    clean_data(data_path)
    error("TEMP @ 579")
    processing_path = os.path.join(PROCESSING_DIR, session_num)
    create_processing_path(processing_path)
    print("\nIdentifying arguments in Session", session_num + "...")
    for deliberation in os.listdir(data_path):
      path = os.path.join(data_path, deliberation)
      if path.endswith('xlsx') or path.endswith('csv'):
        if not IS_SKIP_DATA_PROCESS: create_args_sheet(path, processing_path)
        new_path = os.path.join(processing_path, deliberation)
        formatted_args = extract_args(new_path, deliberation)
        all_args_indexed.update(formatted_args)
        all_args += all_args_indexed[deliberation]
    print("Finished identifying arguments in Session", session_num)

    # sampling 400 arguments for topic extraction
    sampled_args = random.sample(all_args, 400)

    # Generate primary topic [0] and policies [1] - [7]
    topics = extract_topics(sampled_args)
    
    # Generate shorthand variables for each policy and load them into a JSON class
    policy_variables = generate_policy_variables(topics)
    json = build_JSON_class(policy_variables)

    # classify all arguments in Excel files
    #arg_sort(all_args_indexed, topics)
    
    # running inference
    results_path = os.path.join(RESULTS_DIR, session_num)
    argument_analysis_prompt = build_argument_analysis_prompt(topics, policy_variables) if json else error("Cannot generate API prompt: No JSON class")
    arg_inference(all_args_indexed, results_path, argument_analysis_prompt, json)
    delibs = [os.path.join(results_path, csv) for csv in os.listdir(results_path) if csv.endswith(".csv")]

    # running post-evaluation
    cumulative_df = get_metric_sums(delibs, policy_variables[0])
    get_metric_dist(cumulative_df, results_path)
    print("Exiting main() of Session", session_num)

# Code starts here
if __name__ == '__main__':
    print("Program Started")
    # If all sessions should be analyzed
    if IS_ANALYZE_ALL_SESSIONS:
      # Iterate through all sessions from 1 through TOTAL_SESSIONS
      for session in range(TOTAL_SESSIONS):
        session_num = str(session + 1)
        if session_num == '1' or session_num == '2': # Skip 1 - 1 is analyzed
           continue
        main()
    # Else, analyze only Session number 'session_num'.
    else: main()
    print("Program Finished")

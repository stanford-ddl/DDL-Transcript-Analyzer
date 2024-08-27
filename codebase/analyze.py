import os
import sys
import pandas as pd
import random
import ast
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Alignment
import re
# test
from codebase import config, util
from codebase.eval import get_metric_sums, get_metric_dist
from codebase.config import RESULTS_DIR, PROCESSING_DIR

# Given a Worksheet,
# wrap all text cells.
def wrap_text(ws):
   for row in ws.iter_rows():
      for cell in row:
         cell.alignment = Alignment(wrap_text=True)

# Given a deliberation,
# fill in the TRUE/FALSE ending columns.
def label_policy_variables(ws):
   FIRST_SORTED_COL = int(7)
   LAST_SORTED_COL = int(22)
   SORTED_LENGTH = int(LAST_SORTED_COL - FIRST_SORTED_COL + 1)
   for row in range(2, ws.max_row + 1):
      for col in range(FIRST_SORTED_COL, LAST_SORTED_COL + 1):
         cell = ws.cell(row=row, column=col)
         if isinstance(cell.value, str) and cell.value.strip():
            ws.cell(row=row, column=col + SORTED_LENGTH, value="True")
         else: ws.cell(row=row, column=col + SORTED_LENGTH, value="False")

# adds the classified argument to the Excel sheet
def add_arg_result(int_response, line, ws, arg):
    # Calculate the actual row number (add 1 because Excel is 1-indexed)
    actual_row = line + 1

    # Get the current cell value
    current_value = ws.cell(row=actual_row, column=int_response).value

    # If there's already content, append the new arg. Otherwise, use the new arg.
    if current_value:
        new_value = f"{current_value} {arg}"
    else:
        new_value = arg

    # Update the cell with the new value
    ws.cell(row=actual_row, column=int_response, value=new_value)

    if config.is_debug: print(f"(DEBUG) Added '{arg}' to row {actual_row}, column {int_response}")

#cleans up output repsponse by stopping before first newline character
def response_clean(response):
  newline_index = response.find('\n')
  if newline_index != -1:
    response = response[:newline_index]
  backslash_index = response.find('\\')
  # If a backslash is found, return the substring up to the backslash
  if backslash_index != -1:
    response = response[:backslash_index]
  frontslash_index = response.find('/')
  # If a backslash is found, return the substring up to the backslash
  if frontslash_index != -1:
    response = response[:frontslash_index]
  response = response.replace("'", "").replace('"', "").replace('`', "").replace(" ", "")
  return response

# Upades excel file with "for", "against", "other", and "not relevant" columns
def add_policy_columns(ws, policy_variables):
    # Start adding new columns from column 7
    column = 7
    label = ""
    # Do this twice
    for _ in range(2):
      # Add columns for each policy (up to 7)
      for i in range(len(policy_variables)):
          ws.cell(row=1, column=column, value=policy_variables[i] + label)
          column += 1
      label = " (bool)"

def single_deliberation_analysis(deliberation, results_path, all_args_indexed, policy_variables, prompt):
    new_filename = "EVALUATED" + deliberation.replace("xlsx", "csv")
    new_filepath = os.path.join(results_path, new_filename)

    if os.path.exists(new_filepath):
      print("Skipped", deliberation, "because it has previously been analyzed")
      return

    args = all_args_indexed[deliberation]
    path = os.path.join(PROCESSING_DIR, config.session_num, deliberation)

    df = pd.DataFrame(pd.read_excel(path))
    wb = load_workbook(path)
    ws = wb.worksheets[0]

    # add columns to Excel sheets to prepare for classification
    add_policy_columns(ws, policy_variables)

    # loop over a deliberation's arguments
    for arg_group in args:
      process_arg_group(arg_group, prompt, ws)
    
    label_policy_variables(ws)
    wrap_text(ws)

    df = pd.DataFrame(ws.values)
    df.to_csv(new_filepath, index=False, header=False)

# classifies all arguments in all deliberations in session_num based on the generated topics
# note: time-expensive
def arg_sort(all_args_indexed, topics, policy_variables, results_path, transcript_progress_bar, transcript_progress_text, num_transcripts, root):
  print("\nAnalyzing Session", config.session_num, "deliberations...")

  prompt = f"""
You are an annotator categorizing arguments about {topics[0]}. 
If the argument favors {topics[1]}, return "7". If against, return "8".
If the argument favors {topics[2]}, return "9". If against, return "10".
If the argument favors {topics[3]}, return "11". If against, return "12".
If the argument favors {topics[4]}, return "13". If against, return "14".
If the argument favors {topics[5]}, return "15". If against, return "16".
If the argument favors {topics[6]}, return "17". If against, return "18".
If the argument favors {topics[7]}, return "19". If against, return "20".
If relevant but not covered, return "21". If not relevant, return "22".
Return only the number, no extra text or spaces.
"""
  if config.is_debug: print("\n(DEBUG) arg_sort prompt:" + prompt)

  deliberations_analyzed = 0
  # looping over all deliberations
  for deliberation in all_args_indexed.keys():
    single_deliberation_analysis(deliberation, results_path, all_args_indexed, policy_variables, prompt)

    deliberations_analyzed += 1
    transcript_progress_bar['value'] += 100 / num_transcripts
    transcript_progress_text.config(text=f"{deliberations_analyzed}/{num_transcripts}")
    root.title(f"Session {config.session_num} - {(((deliberations_analyzed + (num_transcripts * 2)) / (num_transcripts * 3)) * 100):.2f}% Complete")
    
    print("Analyzed", deliberation)
  print("Finished analyzing deliberations in Session", config.session_num)

# this function takes in a group of arguments, the prompt, and a ws.  It catagorizes every arg in the group
# as for or against one of the topics listed in the prompt and adds it to the appropriate column in the ws.
def process_arg_group(arg_group, prompt, ws):
   line = arg_group[1] # zero based indexed (eg 9 means line 10 of the Excel sheet)
   arg_group_parsed = arg_group[0].split(".")
   for arg in arg_group_parsed:
      if arg == "": # skip empty strings
         continue
      response = util.simple_llm_call(prompt, arg)
      response = response_clean(response)
      counter = 0
      if config.is_debug: print("\n(DEBUG) RESPONSE:", response)
      while response not in [str(i) for i in range(7, 23)]:
        response = util.simple_llm_call(prompt, arg)
        response = response_clean(response)
        counter += 1
        if config.is_debug: print("\n(DEBUG) RETRY:", str(counter))
        if counter >= 5: # couter to prevent infinite loops
           response = "22"
           break
      int_response = int(response)
      add_arg_result(int_response, line, ws, arg)


# Given a list of policies and their corresponding variables,
# generate and save a key.
def generate_key(topics, policy_variables, results_path):
   KEY_NAME = "KEY_Session_" + config.session_num + ".txt"
   key_path = os.path.join(results_path, "metrics", KEY_NAME)

   key_text = "Session " + config.session_num + " Key"
   time = datetime.now().strftime("%Y-%m-%d at %H:%M:%S")
   key_text += "\nGenerated on: " + time
   key_text += "\n\nPrimary Topic: " + topics[0]
   key_text += "\n\nPolicy Key:"
   for i in range(len(policy_variables)):
      key_text += "\n* " + policy_variables[i] + " = " + topics[i+1]
   
   if config.is_debug: print("\n(DEBUG) key_text:\n" + key_text)
   
   with open(key_path, 'w') as key:
      key.write(key_text)

# Given a list of policies,
# return a list of variables that can be used as shorthand for each policy.
def generate_policy_variables(topics, attempts = 0):
   NUM_VARIABLES = str(len(topics) - 1)
   PRIMARY_TOPIC = topics[0]

   policy_list = []
   for i in range(int(NUM_VARIABLES)):
      policy_list.append(topics[i + 1])

   if attempts == 0: print("\nGenerating variables for the", NUM_VARIABLES, "policies regarding", PRIMARY_TOPIC + "...")
   
   prompt = """This is a list of policies representing arguments made in a deliberation about """ + PRIMARY_TOPIC + """.
   Your job is to create variable names for each policy in a Python list of strings, with each string being a variable name.
   Your response will be """ + NUM_VARIABLES + """ strings.
   The variable names should be short.
   The variable names should use camel case style.
   The variable names should encapsulate the core of the policy.
   Here is an example of what you might return:
   ["VariableOne", "VariableTwo",...]
   This program will CRASH if your response fails to conform to these guidelines.
   Other code ran for several hours before you were given this task.
   If this program crashes, we will need to restart which takes several hours.
   Please take your time and ensure ALL of these instructions are followed.
   Do NOT number the list.
   Do NOT indent in your response.
   Do NOT include a newline character in your response.
   Do NOT include anything in the list other than the variable names.
   Your response is ONLY a single list of variable names.
   Once again, your response is ONLY a single list.
   Your response is ONLY one line.
   One again, the variable names must be SHORT
   The Python list you return should contain EXACTLY """ + NUM_VARIABLES + """ strings.
   Thank you!"""
   response = util.simple_llm_call(prompt, policy_list)
   if config.is_debug: print("\n(DEBUG) Raw response:", response + "\n")
   response_list = ast.literal_eval(response.strip())
   # If invalid response, try again
   if len(response_list) != int(NUM_VARIABLES) or type(response_list) != list:
     attempts += 1
     print("Failed attempt #" + str(attempts))
     # If failed too many times, give up
     if attempts >= 5:
       error("Failed to generate variables for each policy.\nThe AI is not cooperating.\nRerun the program, and it should (hopefully) work after a few tries.")
     # Else, try again
     print("Retrying...")
     return generate_policy_variables(topics, attempts)
   # Else, success!
   print_list(response_list, "Variable")

   # Generate results, metrics, and key
   results_path = os.path.join(RESULTS_DIR, config.session_num)
   generate_key(topics, response_list, results_path)
   
   variable_list = []
   for i in range(len(response_list)):
      variable_list.append("FOR: " + response_list[i])
      variable_list.append("AGAINST: " + response_list[i])
   variable_list.append("other")
   variable_list.append("notRelevant")
   if config.is_debug:
      print("\n(DEBUG) variable_list:")
      print_list(variable_list)
   return variable_list

# Called when an error occurs
def error(reason = "No reason provided"):
   print("\n\nERROR:", reason)
   print("\nProgram Terminated")
   sys.exit()

# Given a subset of arguments,
# return a list of topics that most arguments deliberate on.
def extract_topics(sampled_args, attempts = 0):
   NUM_TOPICS = '7'
   if attempts == 0: print("\nGenerating", NUM_TOPICS, "topics from Session", config.session_num + "...")

   prompt = """This is a list of arguments presented in a deliberation. Your job is to identify the single, primary topic deliberated on and write
   """ + NUM_TOPICS + """ distinct policies regarding the primary topic in a Python list of strings, with each string being a policy.
   The first string in the list is the primary topic, and every other string is a policy.
   It should be possible to define every argument initially provided as being an argument "for" or "against" at least one of the policies you generate.
   The exceptions to this are the very small number of arguments that are unrelated to the primary topic and all policies.
   Here is an example of what you might return if the primary topic was 'Ranked choice voting':
   ["Ranked choice voting", "Implement ranked choice voting as an alternative method both to elected officials and representatives at all levels", "Use proportional representatives to elect elected officials"...]
   The policies in the list should be somewhat distinct from each other.
   You should NOT have broad policies, such as the advantages and disadvantages of the primary topic. Rather,
   you should instead find nuanced groups of arguments that may be present on either side of the discussion.
   Every policy should be a feasable, actionable change. Here is an example of what a policy should NOT be:
   "Limit the influence of political parties in elections" This policy is BAD because it is too vague and not actionable;
   this BAD policy could not feasibly be turned into a law.
   The following non-case-sensitive words are BANNED and should NOT be included in your response:
   "and", "but", "or", "by", "Promote", "Encourage", "Ensure", "Explore", "Maintain".
   This program will CRASH if your response fails to conform to these guidelines.
   Other code ran for several hours before you were given this task.
   If this program crashes, we will need to restart which takes several hours.
   Please take your time and ensure ALL of these instructions are followed.
   Do NOT write "Primary topic:" in your response.
   Do NOT number the list.
   Do NOT indent in your response.
   Do NOT include a newline character in your response.
   Do NOT include anything in the list other than the primary topic and the policies themselves.
   Your response is ONLY a single list of the primary topic and the policies.
   Once again, your response is ONLY a single list.
   Your response is ONLY one line.
   VERY IMPORTANT: EVERY argument initially provided to you should be an argument "for" or "against" at least one of the policies you generate.
   Once again, EVERY argument should be relevant to one of the policies you return.
   Please, the Python list you return should contain EXACTLY """ + str(int(NUM_TOPICS)+1) + """ strings.
   Thank you!"""
   response = util.simple_llm_call(prompt, sampled_args)
   if config.is_debug: print("\n(DEBUG) Raw response:", response + "\n")
   topic_list = ast.literal_eval(response.strip())
   # If invalid response, try again
   if len(topic_list) != (int(NUM_TOPICS) + 1) or type(topic_list) != list:
     attempts += 1
     print("Failed attempt #" + str(attempts))
     # If failed too many times, give up
     if attempts >= 5:
       error("Failed to generate a primary topic and policies.\nThe AI is not cooperating.\nRerun the program, and it should (hopefully) work after a few tries.")
     # Else, try again
     print("Retrying...")
     return extract_topics(sampled_args, attempts)
   # Else, success!
   print_topics(topic_list)
   return topic_list

# Given a list and (optionally) a header,
# print it
def print_list(list, header = ""):
   for i in range(len(list)):
      print(header, str(i + 1) + ":", list[i])

# Given a list of topics,
# print them.
def print_topics(topic_list):
  print("Primary Topic:", topic_list[0])
  for i in range(1, len(topic_list)):
    print("Policy", str(i) + ":", topic_list[i])

# Given a key,
# read it and set topics[] and policy_variables[]
def read_key(key, topics, policy_variables):
   print("\nReading pre-existing", key, "key...", end=" ")
   with open(key, 'r') as file:
      text = file.read()
   
   # Get primary topic
   primary_topic = re.search(r'Primary Topic:\s*(.*)', text)
   if primary_topic:
      topics.append(primary_topic.group(1).strip())
   
   # Get policies and variables
   policies = re.findall(r'\* (\w+)\s*=\s*(.*)', text)
   for variable, policy in policies:
      policy_variables.append("FOR: " + variable)
      policy_variables.append("AGAINST: " + variable)
      topics.append(policy)
   policy_variables.append("other")
   policy_variables.append("notRelevant")

   if config.is_debug:
      print()
      print("\n(DEBUG) topics from key:")
      print_topics(topics)
      print("\n(DEBUG) policy_variables from key:")
      print_list(policy_variables, "Variable")
   else: print("Done")

# Given a path for a 'results' folder,
# create that folder and a 'metrics' subfolder if needed.
def create_results_path(results_path):
  print("\nCreating Session", config.session_num, "results and metrics folders...", end=" ")
  # Creates the required results folder if it does not already exist
  os.makedirs(results_path, exist_ok=True)

  # Creates the required metrics folder if it does not already exist
  metrics_path = os.path.join(results_path, "metrics")
  os.makedirs(metrics_path, exist_ok=True)
  print("Done")

# Generate topics[] and policy_variables[].
# If there is a key, read it;
# otherwise, generate it.
def generate_policy_data(sampled_args, topics, policy_variables, results_path):
  create_results_path(results_path)

  # Check for a pre-existing key
  KEY_NAME = "KEY_Session_" + config.session_num + ".txt"
  KEY_PATH = os.path.join(results_path, "metrics", KEY_NAME)
  if os.path.exists(KEY_PATH):
    read_key(KEY_PATH, topics, policy_variables)
    return
  
  # Generate primary topic [0] and policies [1] - [7]
  topics_list = extract_topics(sampled_args)
  for i in range(len(topics_list)):
     topics.append(topics_list[i])
  
  # Generate shorthand variables for each policy and load them into a JSON class
  policy_variables_list = generate_policy_variables(topics)
  for i in range(len(policy_variables_list)):
     policy_variables.append(policy_variables_list[i])

# Code starts here
# Given a results folder and arguments,
# analyze all of the arguments and generate results
def analyze_processed_data(all_args_indexed, all_args, transcript_progress_bar, transcript_progress_text, num_transcripts, root):
  results_path = os.path.join(RESULTS_DIR, config.session_num)
  topics = []
  policy_variables = []
  try:
     sampled_args = random.sample(all_args, 400) # sampling 400 arguments for policy generation
  except ValueError:
     error("There are not enough arguments in this session for a proper analysis!\nGo to analyze_processed_data() in analyze.py if this must be overwritten.")
  generate_policy_data(sampled_args, topics, policy_variables, results_path) # generate or read topics[] and policy_variables[]

  # Classify all arguments in Excel files
  arg_sort(all_args_indexed, topics, policy_variables, results_path, transcript_progress_bar, transcript_progress_text, num_transcripts, root)

  # Generate Metrics
  delibs = [os.path.join(results_path, csv) for csv in os.listdir(results_path) if csv.endswith(".csv")]
  cumulative_df = get_metric_sums(delibs, policy_variables[0])
  get_metric_dist(cumulative_df, results_path)
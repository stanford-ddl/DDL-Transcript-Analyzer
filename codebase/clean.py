import os
import sys
import pandas as pd
from openpyxl import load_workbook

from codebase import config

# Called when an error occurs
def error(reason = "No reason provided"):
   print("\n\nERROR:", reason)
   print("\nProgram Terminated\n")
   sys.exit()

# Given a deliberation,
# convert it from a CSV file into a XLSX file if needed.
def csv_to_xlsx(data_path, deliberation):
  path = os.path.join(data_path, deliberation)
  if path.endswith('numbers'): error("""".numbers" files are not supported. Please manually convert them to ".csv" files using Apple's Numbers Application.\n1. Open the .numbers file in Apple Numbers.\n2. Go to File -> Export To -> CSV.\n3. Save the file as a ".csv".""")
  if path.endswith('csv'):
    df = pd.read_csv(path, header=1)

    # Convert Headers
    headers = pd.DataFrame([df.columns.tolist()], columns=df.columns[:6]) # Get column headers
    df = pd.concat([headers, df]) # Add column headers to the top of the df

    # Convert Room ID
    with open(path, 'r') as f:
      room_ID_list = f.readline().strip().split(',') # Get a list of first row values
    room_ID_row = pd.DataFrame([room_ID_list], columns=df.columns[:2]) # Convert to a df row
    room_ID_row.loc[0, "userId"] = pd.to_numeric(room_ID_row.loc[0, "userId"]) # Convert roomID number from text to a numeric value
    df = pd.concat([room_ID_row, df]) # Add Room ID row to the top of the df

    # Convert from CSV to XLSX
    new_file = deliberation.replace('csv', 'xlsx')
    new_path = os.path.join(data_path, new_file)
    df.to_excel(new_path, index=False, header=False)

    os.remove(path)

# Given a path to a data folder,
# ensure that it already exists;
# otherwise, create it and error
def check_data_exists(data_path):
   print("\nValidating Session", config.session_num, "data folder...", end=" ")
   if not os.path.exists(data_path):
       os.makedirs(data_path, exist_ok=True)
       print() # overrides end=" " from previous print statement
       error("The session " + config.session_num + " data folder does not exist and will now be created.\nPlease place your data into " + data_path + " and restart the program.\nAlternatively, open config.py and select a different session.")
   print("Done")

# Code starts here
# Given a data folder,
# clean the data in it.
def clean_input_data(data_path, transcript_progress_bar, transcript_progress_text, num_transcripts, root):

   check_data_exists(data_path)
   
   print("\nCleaning Session", config.session_num, "data folder...")

   # Convert all CSV files to XLSX files - throws an error for NUMBERS files
   for deliberation in os.listdir(data_path):
      csv_to_xlsx(data_path, deliberation)
   
   deliberations_cleaned = 0
   # Remove all sheets except for the first one from each Workbook
   for deliberation in os.listdir(data_path):
      path = os.path.join(data_path, deliberation)
      if path.endswith('xlsx'):
        wb = load_workbook(path)
        ws = wb.worksheets[0]
        ws.title = "in"
        for sheet_name in wb.sheetnames:
            if sheet_name != "in":
              wb.remove(wb[sheet_name])
        
        # If a number is incorrectly stored as a string,
        # convert it into a number.
        for row in ws.iter_rows():
           for cell in row:
              if cell.value and isinstance(cell.value, str) and cell.value.isdigit():
                 cell.value = int(cell.value)

        # Merge consecutive rows with the same speaker and time
        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row_idx > 2:
                prev_row = ws[row_idx - 1]
                if row[0] == prev_row[0].value and row[1] == prev_row[1].value:  # Same speaker and time
                    prev_row[2].value += " " + (row[2] or "")  # Concatenate text
                    rows_to_delete.append(row_idx)

        # Delete merged rows
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        # Remove rows with empty text
        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[2]:  # Check if text column is empty
                rows_to_delete.append(row_idx)

        # Delete rows with empty text
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        wb.save(path)

        deliberations_cleaned += 1
        transcript_progress_bar['value'] += 100 / num_transcripts
        transcript_progress_text.config(text=f"{deliberations_cleaned}/{num_transcripts}")
        root.title(f"Session {config.session_num} - {((deliberations_cleaned / (num_transcripts * 3)) * 100):.2f}% Complete")

        print("Cleaned", deliberation)
   print("Finished cleaning Session", config.session_num, "data folder")